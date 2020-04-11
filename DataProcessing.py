#import DataImport
import datetime
from pathlib import Path
import pandas as pd
import openpyxl as pyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Color, Font, PatternFill, colors
from openpyxl.styles.borders import Border, Side
from xlsxwriter.utility import xl_col_to_name, xl_rowcol_to_cell
from py_linq import Enumerable
import xlwings as xw
import xlrd
import os
class processData:
    def __init__(self):
        self
    
    def SalesDataProcess(self,SalesData,wb,StartDate,SalesRptXlintoOutPut,entry):
       
        wsSales_Reports  = wb['Sales Reports']        
        Sales_ReportsMaxCol = maxCol(wsSales_Reports,9,1)
        Sales_ReportsMaxRow = maxRow(wsSales_Reports,9,1)

        # Add two columns 
        wsSales_Reports.insert_cols(Sales_ReportsMaxCol)
        wsSales_Reports.insert_cols(Sales_ReportsMaxCol)


        for i in range(10,Sales_ReportsMaxRow + 1):
            if SalesData.loc[SalesData['Item SKU Code'] == wsSales_Reports.cell(i,2).value,'Sale Order Status'].count() == 1:
                wsSales_Reports.cell(i,Sales_ReportsMaxCol).value = SalesData.loc[SalesData['Item SKU Code'] == wsSales_Reports.cell(i,2).value,'Sale Order Status'].values[0]
            else:
                wsSales_Reports.cell(i,Sales_ReportsMaxCol).value = 0

        formulaSumSQ =  '=SUM(' + xl_col_to_name(Sales_ReportsMaxCol -1,True) + str(10) + ':' + xl_col_to_name(Sales_ReportsMaxCol - 1,True) + str(Sales_ReportsMaxRow) + ")"
        formulaSumAP = '=SUM(' + xl_col_to_name(Sales_ReportsMaxCol ,True) + str(10) + ':' + xl_col_to_name(Sales_ReportsMaxCol,True) + str(Sales_ReportsMaxRow) + ")"
        FormulaBQ = '=SUM(' + xl_col_to_name(Sales_ReportsMaxCol + 1,True) + str(10) + ':' + xl_col_to_name(Sales_ReportsMaxCol + 1,True) + str(Sales_ReportsMaxRow) + ")"

        wsSales_Reports[xl_col_to_name(Sales_ReportsMaxCol -1) + str(Sales_ReportsMaxRow + 1)].value = formulaSumSQ
        wsSales_Reports[xl_col_to_name(Sales_ReportsMaxCol) + str(Sales_ReportsMaxRow + 1)].value = formulaSumAP
        wsSales_Reports[xl_col_to_name(Sales_ReportsMaxCol + 1) + str(Sales_ReportsMaxRow + 1)].value = FormulaBQ


        marginPercent = round((1 - wsSales_Reports['E10'].value),2)
        strRange = xl_col_to_name(Sales_ReportsMaxCol)  + '10' + ':' + xl_col_to_name(Sales_ReportsMaxCol) + str(Sales_ReportsMaxRow)

        formulaAP = '=(' + str(marginPercent) + '*$D{0})*' + xl_col_to_name(Sales_ReportsMaxCol -1 ,True) + '{0}'
        for i, cellObj in enumerate(wsSales_Reports[strRange], 10):
            cellObj[0].value = formulaAP.format(i)

        for i in range(10,Sales_ReportsMaxRow + 1):
            wsSales_Reports.cell(i,Sales_ReportsMaxCol + 2).value = "=" + GetFormuleBQ(Sales_ReportsMaxCol + 2,i)[:-1]
        
        blackFill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
        ft_White = Font(color=colors.WHITE)
        ft_WhiteBold = Font(bold=True,color=colors.WHITE)
        

        TempDate = datetime.datetime.strptime(StartDate, "%d/%m/%Y") 
        TempDate2 = datetime.datetime.strftime(TempDate, "%d-%m-%Y")
        TempDate3 = datetime.datetime.strftime(TempDate, "%b")

        _cell1 = wsSales_Reports.cell(9,Sales_ReportsMaxCol)         
        _cell1.font = ft_White       
        _cell1.fill = blackFill
        #_cell1.value = datetime.datetime.strftime(StartDate, "%b") + " " + datetime.datetime.strftime(StartDate, "%d/%m/%Y")
        _cell1.value =  TempDate3 + " " + TempDate2 
        _cell1.alignment = Alignment(wrapText=True)


        _cell2 = wsSales_Reports.cell(9,Sales_ReportsMaxCol + 1)        
        _cell2.font = ft_White
        _cell2.fill = blackFill
        _cell2.value = 'Amount Payable'
        _cell2.alignment = Alignment(wrapText=True)
        

        _cell3 = wsSales_Reports.cell(Sales_ReportsMaxRow + 1,Sales_ReportsMaxCol)
        _cell3.font = ft_WhiteBold
        _cell3.fill = blackFill
        
        _cell4 = wsSales_Reports.cell(Sales_ReportsMaxRow + 1,Sales_ReportsMaxCol + 1)
        _cell4.font = ft_WhiteBold
        _cell4.fill = blackFill

        strRange = xl_col_to_name(Sales_ReportsMaxCol - 1)  + '10' + ':' + xl_col_to_name(Sales_ReportsMaxCol) + str(Sales_ReportsMaxRow + 1)
        
        for row_cells in wsSales_Reports[strRange]:
            for cell in row_cells:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        set_border(wsSales_Reports,strRange)
        
        import ntpath
        testpath =  ntpath.basename(entry)
        wb.save(SalesRptXlintoOutPut.joinpath(testpath))
        

        #wb = xw.Book(app_visible=False))
        strPath = os.path.abspath(str(SalesRptXlintoOutPut.joinpath(testpath)))
        wb = xw.Book(strPath)
        app = xw.apps.active
        app.visible = False
        #xw.visible = False
        sht = wb.sheets['Sales Reports']
        my_values = sht.range('$G10:' + xl_col_to_name(Sales_ReportsMaxCol ,True) + str(Sales_ReportsMaxRow) ).options(ndim=2).value         
        sht.range('$G10').value = my_values

        for i in range(10,Sales_ReportsMaxRow):
            intBalQ = sht.range(xl_col_to_name(Sales_ReportsMaxCol + 1,True) + str(i)).value
            intLastCol = sht.range(xl_col_to_name(Sales_ReportsMaxCol - 1,True) + str(i)).value            
            if(intBalQ < 0):
                sht.range(xl_col_to_name(Sales_ReportsMaxCol - 1,True) + str(i)).value = (int(intLastCol) - abs(intBalQ))

        wb.save()
        #wb.close()
        app.quit()
        

        
        #sht.range('E8').value
        #wb2 = pyxl.load_workbook(filename= SalesRptXlintoOutPut.joinpath(testpath))
        #wsSales_Reports2 = wb2['Sales Reports']   
        #read_fullreport = xlrd.open_workbook(SalesRptXlintoOutPut.joinpath(testpath))
        #read_fullreport_sheet = read_fullreport.sheet_by_index(0)
        #ipAddressofMC = read_fullreport_sheet.cell_value(10, 8)
        #val = wsSales_Reports2['H10'].value
        #xx = ipAddressofMC2


    def StockDataProcess(self,StockData,wb,StartDate,end_date):
        wsStock_Update  = wb['Stock Update']

        wsStock_UpdateMaxCol = maxCol(wsStock_Update,1,1)
        wsStock_UpdateMaxRow = maxRow(wsStock_Update,1,1)

        from datetime import timedelta, date

        #StartDate =  datetime.datetime.strptime(StartDate, "%d/%m/%Y")
        #end_date = datetime.datetime.strptime(end_date, "%d/%m/%Y")

        StartDate = datetime.datetime.strptime(StartDate, "%d/%m/%Y")         

        end_date = datetime.datetime.strptime(end_date, "%d/%m/%Y")
        
        #Adjustment for end date
        end_date = end_date + datetime.timedelta(days=1)

        def daterange(self,StartDate, end_date):
            for n in range(int ((end_date - StartDate).days)):
                yield StartDate + timedelta(n)


        #marks = Enumerable(StockData)
        #passing = marks.where(lambda x: x['Item SkuCode'] == 'CONHSOFQRK001') # results in [50, 80, 90]

        #print(passing)

        col = 1
        col2 = 0
        for single_date in daterange(self,StartDate, end_date):
            Total = 0
            for i in range(2,wsStock_UpdateMaxRow + 1): # Loop through excel data
                if StockData.loc[StockData['Item SkuCode'] == wsStock_Update.cell(i,2).value,'Quantity Received'].count() >= 1:
                    #wsStock_Update.cell(i,wsStock_UpdateMaxCol + col).value = StockData.loc[StockData['Item SkuCode'] == wsStock_Update.cell(i,2).value].values[0]
                    #if datetime.datetime.strptime(StockData.loc[StockData['GRN Date']],'%d-%m-%Y').date() == single_date:
                    df = StockData.loc[(StockData['Item SkuCode'] == wsStock_Update.cell(i,2).value)]
                    df['GRN Date'] = pd.to_datetime(df['GRN Date'])  
                    df = df.set_index(['GRN Date'])                
                    #df['GRN Date'] = df['GRN Date'].dt.date
                    df = df.loc[single_date:single_date]
                    df2 = df.reset_index()
                    if df2.loc[df2['GRN Date'] == single_date,'Quantity Received'].count() == 1:
                                            Total = Total + df2.loc[(df2['GRN Date'] == single_date,'Quantity Received')].values[0]
                    else:
                        Total = Total + 0
                else:
                    Total = Total + 0
            if Total <= 0 :
                continue

            for i in range(2,wsStock_UpdateMaxRow + 1): # Loop through excel data
                if StockData.loc[StockData['Item SkuCode'] == wsStock_Update.cell(i,2).value,'Quantity Received'].count() >= 1:
                    #wsStock_Update.cell(i,wsStock_UpdateMaxCol + col).value = StockData.loc[StockData['Item SkuCode'] == wsStock_Update.cell(i,2).value].values[0]
                    #if datetime.datetime.strptime(StockData.loc[StockData['GRN Date']],'%d-%m-%Y').date() == single_date:
                    df = StockData.loc[(StockData['Item SkuCode'] == wsStock_Update.cell(i,2).value)]
                    df['GRN Date'] = pd.to_datetime(df['GRN Date'])    
                    df = df.set_index(['GRN Date'])                
                    #df['GRN Date'] = df['GRN Date'].dt.date
                    df = df.loc[single_date:single_date]
                    df2 = df.reset_index()
                    if df2.loc[df2['GRN Date'] == single_date,'Quantity Received'].count() == 1:
                                            wsStock_Update.cell(i,wsStock_UpdateMaxCol + col).value = df2.loc[(df2['GRN Date'] == single_date,'Quantity Received')].values[0]
                    else:
                        wsStock_Update.cell(i,wsStock_UpdateMaxCol + col).value = 0
                else:
                    wsStock_Update.cell(i,wsStock_UpdateMaxCol + col).value = 0

#print(df.loc[(df['GRN Date'] == '30-4-2019')])
            FormulaSQ = '=SUM(' + xl_col_to_name(wsStock_UpdateMaxCol + col2,True) + str(2) + ':' + xl_col_to_name(wsStock_UpdateMaxCol + col2,True) + str(wsStock_UpdateMaxRow) + ")"

            wsStock_Update[xl_col_to_name(wsStock_UpdateMaxCol + col2,True) + str(wsStock_UpdateMaxRow + 1)].value = FormulaSQ

            FormulaPB = '=SUM(' + xl_col_to_name(3,True) + str(2) + ':' + xl_col_to_name(3,True) + str(wsStock_UpdateMaxRow) + ")"

            wsStock_Update[xl_col_to_name(3,True) + str(wsStock_UpdateMaxRow + 1)].value = FormulaPB
            
            strRange = xl_col_to_name(3)  + '2' + ':' + xl_col_to_name(3) + str(wsStock_UpdateMaxRow)

            formulaAP = '=SUM(' + xl_col_to_name(4,True) + '{0}' + ':' + xl_col_to_name(wsStock_UpdateMaxCol + col2,True) + '{0}'+ ")"
            for i, cellObj in enumerate(wsStock_Update[strRange], 2):
                cellObj[0].value = formulaAP.format(i)



            ## Formating starts here 
            blackFill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
            ft_White = Font(color=colors.WHITE)
           #ft_WhiteBold = Font(bold=True,color=colors.WHITE)
            ft_BlackBold = Font(bold=True,color=colors.BLACK)

            #DateTemp = datetime.datetime.strptime(single_date, "%d/%m/%Y") 

            _cell1 = wsStock_Update.cell(1,wsStock_UpdateMaxCol + col)        
            _cell1.font = ft_White
            _cell1.fill = blackFill
            _cell1.value = 'Qty Rcvd ' + datetime.datetime.strftime(single_date, "%d/%m/%Y") 
            _cell1.alignment = Alignment(wrapText=True)
            
            #side = Side(border_style='thin', color="FF000000")

            _cell2 = wsStock_Update.cell(wsStock_UpdateMaxRow + 1,wsStock_UpdateMaxCol + col)        
            _cell2.font = ft_BlackBold
            _cell2.alignment = Alignment(horizontal='center', vertical='center')            

            #_cell2.fill = blackFill
            
            strRange = xl_col_to_name(wsStock_UpdateMaxCol + col2,True)  + '2' + ':' + xl_col_to_name(wsStock_UpdateMaxCol + col2) + str(wsStock_UpdateMaxRow + 1)

            for row_cells in wsStock_Update[strRange]:
                for cell in row_cells:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            set_border(wsStock_Update,strRange)
            col = col + 1
            col2 = col2 + 1

        for jj in range(2,wsStock_UpdateMaxRow + 1):
            mxColNow = maxCol(wsStock_Update,1,1)
            for j in range(5,mxColNow):
                formula = xl_col_to_name(j)
                #print(wsStock_Update.cell(1,j + 2).value)
                if mxColNow == 6:
                    cell_ = wsStock_Update.cell(1,j + 1).value

                    if 'Returned' not in cell_:                
                            formula1 = 'E' + str(jj) + '+' + (formula + str(jj))
                    if 'Returned' in cell_:
                            formula1 = 'E' + str(jj) + '-' + (formula + str(jj))
                else :
                    cell_ = wsStock_Update.cell(1,j + 1).value
                              
                    if j == 5 :                                        
                        if 'Returned' not in cell_:                
                            formula1 = 'E' + str(jj) + '+' + (formula + str(jj)) 
                        if 'Returned' in cell_:
                            formula1 = 'E' + str(jj) + '+' + (formula + str(jj)) 

                    elif (j > 5) & (j <  mxColNow - 1):

                        if 'Returned' not in cell_:
                            formula1 = formula1 +  '+' + (formula + str(jj)) 
                        if 'Returned' in cell_:
                            formula1 = formula1 +  '-' + (formula + str(jj)) 
                            
                    elif (j ==  mxColNow -1):
                         if 'Returned' not in cell_:
                            formula1 = formula1 + '+' + (formula + str(jj))
                         if 'Returned' in cell_:
                             formula1 = formula1 + '-' + (formula + str(jj))
            if mxColNow == 5:
                cell_ = wsStock_Update.cell(1,5).value
                    
                if 'Returned' not in cell_:                
                    formula1 = 'E' + str(jj)
                if 'Returned' in cell_:
                    formula1 = '-' + 'E' + str(jj) 

            wsStock_Update.cell(jj,4).value = "=" + formula1




    def NewSkusFinder(self,wb):
        wsSR  = wb['Sales Reports']
        wsStock_UpdateMaxRow = maxRow(wsSR,9,1)
        drSR = lstDataset(wsSR,'B10','B' + str(wsStock_UpdateMaxRow))
        dfwsBrand_Awareness = pd.DataFrame(drSR)
        return dfwsBrand_Awareness


# Read the cell values into a list of lists
def lstDataset(ws,strStart,strEnd):
    data_rows = []
    for row in ws[strStart:strEnd]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    return data_rows


def maxCol(ws,intStartRow,intStartCol):
    maxcol = intStartCol - 1 
    for i in range(intStartCol,16384):
        if ws.cell(intStartRow,i).value == None:
            break
        #print(ws.cell(intStartRow,i).value)
        maxcol = maxcol + 1    
    return maxcol

def maxRow(ws,intStartRow,intStartCol):
    maxrow = intStartRow -1
    for i in range(intStartRow,1048576):
        if ws.cell(i,intStartCol).value == None:            
            break
        #print(ws.cell(i,intStartCol).value)
        maxrow = maxrow + 1    
    return maxrow

def GetFormuleBQ(intMaxCol,inRow):    
    intcount = 1
    for i in range(7,intMaxCol):
        if intcount == 3:
            intcount = 1
        
        if intcount == 1:
            formula = xl_col_to_name(i - 1)
            if i == 7:
                formula1 = 'F' + str(inRow) + '-' + (formula + str(inRow)) +  '-' 
            elif i > 7 & i != intMaxCol:    
                formula1 = formula1 + (formula + str(inRow) + '-')        
            intcount = intcount + 1
        else:
            intcount = intcount + 1
    return formula1


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell        
        for pos_x, cell in enumerate(cells):                       
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom                
            )
            border.left = side
            border.right = side
            border.top = side
            border.bottom = side           
            #if pos_x == 0:
            #    border.left = side
            #if pos_x == max_x:
            #    border.right = side
            #if pos_y == 0:
            #    border.top = side
            #if pos_y == max_y:
            #    border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border
