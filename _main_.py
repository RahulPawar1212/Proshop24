import DataImport
import DataProcessing
import openpyxl as pyxl
from pathlib import Path
import dateparser
import datetime
import os
import shutil
import glob
#from tkinter import *
from tkinter import filedialog,messagebox,Entry,Button,StringVar,Label,Tk
import tkinter as tk 
import time
#from xlsxwriter.utility import xl_rowcol_to_cell,xl_col_to_name
#from openpyxl.styles import Fill,Font,Color
#from openpyxl.styles import colors

class StartProcess:
    def __init__(self):
        root = tk.Tk()        

        self.RunStockProcess  = False
        self.RunSalesProcess = False

        def SelectStockProcess():
            self.RunStockProcess = True            
            root.quit()
            root.withdraw()

        def SelectSalesProcess():
            self.RunSalesProcess = True           
            root.quit()
            root.withdraw()

        #SelectPFrom = Tk()
        root.title("Select Process")
        root.geometry('250x100')

        StockButton = Button(root, text="Stock Update", fg="Black", bg="Green",command=SelectStockProcess) 
        StockButton.grid(row=7, column=2)
      
        SalesButton = Button(root, text="Sales Process", fg="Black", bg="Green",command=SelectSalesProcess) 
        SalesButton.grid(row=8, column=2)
      
        root.mainloop()
     

        #filedialog = filedialog()
        if self.RunStockProcess == True:
            ##
            print("Select GRN file.")
            self.GrnPath  = filedialog.askopenfilename(initialdir="/",
                                                    title = "Select GRN File",
                                                    filetypes = (("Excel Files","*.csv"),("All Files","*.*")))

            print("GRN file path : " + self.GrnPath)
        
        ##
        if self.RunSalesProcess == True:
            print("Select sales data file.")
            self.SalesDataPath  = filedialog.askopenfilename(initialdir="/",
                                                    title = "Select sales data path",
                                                    filetypes = (("Excel Files","*.csv"),("All Files","*.*")))
            
            print("Sales Path : " + self.SalesDataPath)
        
        ##
        print("Select sales reports file's folder")
        self.SalesReporFolderPath = filedialog.askdirectory(mustexist=True,title = "Select sales reports file's folder")
        print("Sales reports folder : " + self.SalesReporFolderPath)

        root.withdraw()

        main1 = Tk()
        main1.title("Select Dates")
        main1.geometry('250x100')
     
        def close_window():
            self.getStartDate_ = getStartDate.get()
            self.getEndDate_ = getEndDate.get() 
            main1.quit()
            root.quit()
            main1.withdraw()    
        
        lbl1 = Label(main1, text="Start Date", bg="light green") 
        lbl1.grid(row=0, column=0)
            
        lbl2 = Label(main1, text="End Date", bg="light green") 
        lbl2.grid(row=2, column=0)
        
        getStartDate = Entry(main1) 
        getEndDate = Entry(main1) 

        getStartDate.grid(row=0, column=2, ipadx="30") 
        getEndDate.grid(row=2, column=2, ipadx="30") 

        
        submit = Button(main1, text="Submit", fg="Black", bg="Red",command=close_window) 
        submit.grid(row=8, column=2)

        main1.mainloop()

    def main(self):
            
        #submit.quit
        #root.destroy()
        
        ## Execution
        if self.RunStockProcess == True: 
            GrnDataPath = Path(self.GrnPath)
        
        if self.RunSalesProcess == True:
            SalesDataPath = Path(self.SalesDataPath)
        
        SalesReporExcel = Path(self.SalesReporFolderPath)
        #print("Enter start date")
        #getStartDate = input()

        #print("Enter end date")
        #getEndDate = input()
        #test = self.getStartDate
        
        StartDate = self.getStartDate_
        EndDate = self.getEndDate_
        #GrnDataPath = Path(r"C:\Users\rahul.pawar\Desktop\Proshop\Goods Received Notes (Stock Intake).csv")
        #SalesDataPath = Path(r"C:\Users\rahul.pawar\Desktop\Proshop\Sales Export From Backend.csv")
        #SalesReporExcel = Path(r"C:\Users\rahul.pawar\Desktop\Proshop")
        
        #StartDate = dateparser.parse('27-8-2019')
        #EndDate = dateparser.parse('29-8-2019')

        # Get string path from path object -- Not useful if selecting file from dialog
        strSalesReporExcel = os.path.abspath(str(SalesReporExcel))
        
        #************************** Import Data *******************************
        Di = DataImport.dataImport()

        if self.RunStockProcess == True:
            GRNdata = Di.getGRNData(GrnDataPath,StartDate,EndDate)
        
        #print(GRNdata)
        if self.RunSalesProcess == True:
            SalesData = Di.getsalesData(SalesDataPath,StartDate,EndDate)
        #*****************************************************************
#
        if self.RunStockProcess == True:
            SalesRptXlintoOutPut = SalesReporExcel
            SalesRptXlintoOutPut = SalesRptXlintoOutPut.joinpath('Stock Updates Output')

            # Create folder if not exists / delete if exists and recreate folder
            if not os.path.exists(SalesRptXlintoOutPut):
                os.makedirs(SalesRptXlintoOutPut)
            elif os.path.exists(SalesRptXlintoOutPut):
                shutil.rmtree(SalesRptXlintoOutPut)
                os.makedirs(SalesRptXlintoOutPut)

            import ntpath
            strSalesReporExcel
            
            # Get file names from path along with path
            files = glob.glob(strSalesReporExcel + '\\' + '*.xlsx')
            for entry in files:    
                # Sales report excel
                #wb = pyxl.load_workbook(filename= SalesReporExcel.joinpath ('Sample Report 2.xlsx') ,read_only=False)
                wb = pyxl.load_workbook(filename= entry ,read_only=False)

                Dp = DataProcessing.processData()
                Dp.StockDataProcess(GRNdata,wb,StartDate,EndDate)   
                
                testpath =  ntpath.basename(entry)
                wb.save(SalesRptXlintoOutPut.joinpath(testpath))

#       

        if self.RunSalesProcess == True:
            strSalesReporExcel2 = os.path.normpath(strSalesReporExcel + os.sep + os.pardir)
            #SalesRptXlintoOutPut2 = SalesReporExcel
            SalesRptXlintoOutPut2 = Path(strSalesReporExcel2)
            SalesRptXlintoOutPut2 = SalesRptXlintoOutPut2.joinpath('Final Output')

            # Create folder if not exists / delete if exists and recreate folder
            if not os.path.exists(SalesRptXlintoOutPut2):
                os.makedirs(SalesRptXlintoOutPut2)
            elif os.path.exists(SalesRptXlintoOutPut2):
                shutil.rmtree(SalesRptXlintoOutPut2)
                os.makedirs(SalesRptXlintoOutPut2)

            import ntpath
            #strSalesReporExcel
            
            # Get file names from path along with path
            files = glob.glob(strSalesReporExcel2 + '\\Stock Updates Output\\' + '*.xlsx')
            for entry in files:    
                # Sales report excel
                #wb = pyxl.load_workbook(filename= SalesReporExcel.joinpath ('Sample Report 2.xlsx') ,read_only=False)
                wb = pyxl.load_workbook(filename= entry ,read_only=False)

                Dp = DataProcessing.processData()                
                Dp.SalesDataProcess(SalesData,wb,StartDate)                                                

                testpath =  ntpath.basename(entry)
                wb.save(SalesRptXlintoOutPut2.joinpath(testpath))


        print("Report Generated....")
        #time.sleep(int(4))

        messagebox.showinfo("Message.", "Reports generated")
        #Testing
        #print(Sales_ReportsMaxCol)
        #print(Sales_ReportsMaxRow)
        #print(wsStock_UpdateMaxCol)
        #print(wsStock_UpdateMaxRow)
